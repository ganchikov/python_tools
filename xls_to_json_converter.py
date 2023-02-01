from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
 
# константы - названия полей в файле эксель
CONST_COL_NAME_CULTURE = 'Культура'
CONST_COL_NAME_FIELD = 'Поле'
CONST_COL_NAME_HARVEST_FACT = 'Валовый сбор, тн'
CONST_FILENAME = 'data_300123'

wb = load_workbook(filename=CONST_FILENAME+'.xlsx')
ws = wb.active
 
# выходной JSON 
output = {
    "crops": [],
    "fields": []
}
 
last_column = len(list(ws.columns))
last_row = len(list(ws.rows))
 
for row in range(2, last_row + 1):
    crop_key = ""
    field_key = ""
    gross_harvest = 0
    crop_key_found = False
    
    for column in range(1, last_column + 1):
        column_letter = get_column_letter(column)
        column_name = ws[column_letter + str(1)].value

        # если колонка = 'Культура'
        if column_name == CONST_COL_NAME_CULTURE:
            crop_key = ws[column_letter + str(row)].value
            for crop_item in output["crops"]:
                if crop_item['key'] == crop_key : crop_key_found = True
            
        # если колонка = 'Поле'
        if column_name == CONST_COL_NAME_FIELD:
            field_key = ws[column_letter + str(row)].value

        # если конока = 'Факт уборки, га'
        if column_name == CONST_COL_NAME_HARVEST_FACT:
            gross_harvest = ws[column_letter + str(row)].value
                
    # если в выходном JSON нет такой культуры, то добавляем запись 'crop'
    if crop_key_found == False: 
        output["crops"].append(
            {
                "key": crop_key, 
                "base_humidity": 0, 
                "compatible_storage_types": [ ],
                "compatible_silo_types": [ ],
                "attributes": [
                    {
                    "key": '',
                    "value": ''
                    }
                ]
            }
        )

    # добавляем в выходной json запись 'field'
    output["fields"].append(
            {
            "key": field_key,
            "crop_key": crop_key,
            "capacity_forecast": [],
            "humidity_forecast": [],
            "amount": gross_harvest,
            "attributes": []
        }
    )
 
json_string = json.dumps(output, sort_keys=True, indent=4, ensure_ascii=False)
with open(CONST_FILENAME+'.json', 'w', encoding='utf-8') as f:
    f.write(json_string)   